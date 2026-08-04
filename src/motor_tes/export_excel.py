"""Puente Python -> Excel: exporta la curva calibrada a un libro con rangos con nombre.

Replica el flujo de una mesa: la calibración pesada corre en Python contra los datos
del Banco de la República, y el pricing intradía ocurre en Excel leyendo los parámetros
ya calibrados. Las UDFs de VBA (``excel/vba/*.bas``) consumen el rango ``NSS_PARAMS``
que escribe este módulo, así que el libro funciona en cualquier máquina con Excel, sin
Python instalado.

Hojas que produce
-----------------
``Parametros``
    Los seis parámetros NSS más spot, SOFR y fecha de la curva, con rangos con nombre.
``Nodos``
    Los nodos de mercado usados en la calibración, con su fuente y residual.
``Curva``
    Grilla de plazos con tasa cero cupón, factor de descuento y forward instantánea.
``Forwards``
    Muestra de precios forward USD/COP con sus sensibilidades.
``Validacion``
    Valores de referencia calculados en Python junto a la fórmula VBA equivalente, para
    contrastar ambas implementaciones dentro del libro.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Final

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from motor_tes import config
from motor_tes.curva_nss import (
    NSSParams,
    ResultadoCalibracion,
    duracion_modificada,
    dv01,
    factor_descuento,
    flujos_bono,
    tasa_cero_cupon,
    tasa_forward,
    tasa_forward_instantanea,
    valor_presente,
)
from motor_tes.pricer_forward import comparar_convenciones, pricer_forward

__all__ = [
    "RUTA_LIBRO_BASE",
    "PLAZOS_GRILLA_ANIOS",
    "PLAZOS_FORWARD_DIAS",
    "exportar_libro",
]

#: Libro sin macros que genera este módulo. ``excel/build_excel.py`` lo convierte en
#: ``.xlsm`` incorporando los módulos VBA.
RUTA_LIBRO_BASE: Final[Path] = config.DIR_EXCEL / "motor_tes_forwards.xlsx"

#: Plazos de la hoja ``Curva``, en años.
PLAZOS_GRILLA_ANIOS: Final[tuple[float, ...]] = (
    1 / 365,
    1 / 12,
    0.25,
    0.5,
    0.75,
    1.0,
    1.5,
    2.0,
    3.0,
    4.0,
    5.0,
    6.0,
    7.0,
    8.0,
    9.0,
    10.0,
)

#: Plazos de la hoja ``Forwards``, en días calendario.
PLAZOS_FORWARD_DIAS: Final[tuple[int, ...]] = (30, 60, 90, 180, 270, 360, 540, 720)

#: Bono de ejemplo para la hoja de validación: cupón, años, frecuencia.
BONO_EJEMPLO: Final[tuple[float, float, int]] = (0.1325, 7.0, 1)

_TITULO = Font(bold=True, size=13)
_CABECERA = Font(bold=True, color="FFFFFF")
_RELLENO = PatternFill("solid", fgColor="1F3864")
_ETIQUETA = Font(bold=True)
_BORDE = Border(bottom=Side(style="thin", color="AAAAAA"))


def _escribir_tabla(hoja, df: pd.DataFrame, fila_inicial: int = 1) -> None:
    """Vuelca un ``DataFrame`` con el encabezado formateado."""
    for j, columna in enumerate(df.columns, start=1):
        celda = hoja.cell(row=fila_inicial, column=j, value=str(columna))
        celda.font = _CABECERA
        celda.fill = _RELLENO
        celda.alignment = Alignment(horizontal="center")
        hoja.column_dimensions[get_column_letter(j)].width = max(len(str(columna)) + 2, 14)

    for i, (_, registro) in enumerate(df.iterrows(), start=fila_inicial + 1):
        for j, valor in enumerate(registro, start=1):
            if isinstance(valor, (pd.Timestamp, datetime, date)):
                valor = pd.Timestamp(valor).strftime("%Y-%m-%d")
            elif isinstance(valor, (np.integer, np.floating)):
                valor = valor.item()
            hoja.cell(row=i, column=j, value=valor)


def _hoja_parametros(
    wb: Workbook,
    calibracion: ResultadoCalibracion,
    spot: float,
    sofr: float,
    fecha_curva: date,
) -> None:
    """Escribe la hoja de parámetros y declara los rangos con nombre que lee el VBA."""
    hoja = wb.active
    hoja.title = "Parametros"

    hoja["A1"] = "Motor de Valoración — curva COP calibrada"
    hoja["A1"].font = _TITULO
    hoja["A2"] = (
        "Calibrado en Python contra datos del Banco de la República. "
        "Las UDFs de VBA leen estos valores."
    )

    hoja["A4"] = "Parámetros Nelson-Siegel-Svensson"
    hoja["A4"].font = _ETIQUETA
    etiquetas = ("beta0", "beta1", "beta2", "beta3", "lambda1", "lambda2")
    descripciones = (
        "Nivel de largo plazo",
        "Pendiente (beta0+beta1 = tasa instantánea)",
        "Primera curvatura",
        "Segunda curvatura (0 en Nelson-Siegel puro)",
        "Escala temporal 1, en años",
        "Escala temporal 2, en años",
    )
    vector = calibracion.params.to_array()
    for i, (etiqueta, valor, descripcion) in enumerate(
        zip(etiquetas, vector, descripciones), start=5
    ):
        hoja.cell(row=i, column=1, value=etiqueta).font = _ETIQUETA
        hoja.cell(row=i, column=2, value=float(valor)).number_format = "0.00000000"
        hoja.cell(row=i, column=3, value=descripcion)

    hoja["A12"] = "Mercado"
    hoja["A12"].font = _ETIQUETA
    hoja["A13"] = "TRM spot (COP/USD)"
    hoja["B13"] = float(spot)
    hoja["B13"].number_format = "#,##0.00"
    hoja["A14"] = "SOFR (decimal, simple ACT/360)"
    hoja["B14"] = float(sofr)
    hoja["B14"].number_format = "0.0000%"
    hoja["A15"] = "Fecha de la curva"
    hoja["B15"] = pd.Timestamp(fecha_curva).strftime("%Y-%m-%d")

    hoja["A17"] = "Diagnóstico de la calibración"
    hoja["A17"].font = _ETIQUETA
    diagnostico = {
        "Modelo": "Svensson (6p)" if calibracion.svensson else "Nelson-Siegel (4p)",
        "Nodos": len(calibracion.plazos),
        "Grados de libertad": calibracion.grados_libertad,
        "RMSE (bps)": round(calibracion.rmse_bps, 4),
        "Max |residual| (bps)": round(
            float(np.max(np.abs(calibracion.residuales_bps))), 4
        ),
        "Convergió": "Sí" if calibracion.exito else "No",
    }
    for i, (clave, valor) in enumerate(diagnostico.items(), start=18):
        hoja.cell(row=i, column=1, value=clave)
        hoja.cell(row=i, column=2, value=valor)

    if calibracion.es_interpolacion:
        hoja["A25"] = (
            "AVISO: 0 grados de libertad. El modelo interpola en vez de ajustar, "
            "así que un RMSE nulo no es evidencia de calidad."
        )
        hoja["A25"].font = Font(bold=True, color="C00000")

    hoja.column_dimensions["A"].width = 34
    hoja.column_dimensions["B"].width = 18
    hoja.column_dimensions["C"].width = 46

    for nombre, referencia in (
        ("NSS_PARAMS", "Parametros!$B$5:$B$10"),
        ("TRM_SPOT", "Parametros!$B$13"),
        ("SOFR_ACTUAL", "Parametros!$B$14"),
        ("CURVA_FECHA", "Parametros!$B$15"),
    ):
        wb.defined_names.add(DefinedName(nombre, attr_text=referencia))


def _hoja_nodos(
    wb: Workbook, calibracion: ResultadoCalibracion, nodos: pd.DataFrame
) -> None:
    """Nodos de mercado usados, con el residual que dejó la calibración."""
    hoja = wb.create_sheet("Nodos")
    tabla = pd.DataFrame(
        {
            "plazo_anios": calibracion.plazos,
            "tasa_mercado": calibracion.tasas_mercado,
            "tasa_ajustada": calibracion.tasas_ajustadas,
            "residual_bps": calibracion.residuales_bps,
        }
    )
    if "fuente" in nodos.columns and len(nodos) == len(tabla):
        tabla.insert(1, "fuente", nodos.sort_values("plazo_anios")["fuente"].to_numpy())
    _escribir_tabla(hoja, tabla.round(8))


def _hoja_curva(wb: Workbook, params: NSSParams) -> None:
    """Grilla de la curva: tasa, factor de descuento y forward instantánea."""
    hoja = wb.create_sheet("Curva")
    plazos = np.array(PLAZOS_GRILLA_ANIOS, dtype=float)
    tabla = pd.DataFrame(
        {
            "plazo_anios": plazos,
            "tasa_cero_cupon": np.asarray(tasa_cero_cupon(plazos, params)),
            "factor_descuento": np.asarray(factor_descuento(plazos, params)),
            "forward_instantanea": [
                tasa_forward_instantanea(float(t), params) for t in plazos
            ],
        }
    )
    _escribir_tabla(hoja, tabla.round(10))


def _hoja_forwards(wb: Workbook, params: NSSParams, spot: float, sofr: float) -> None:
    """Muestra de forwards USD/COP con sus sensibilidades."""
    hoja = wb.create_sheet("Forwards")
    filas = []
    for dias in PLAZOS_FORWARD_DIAS:
        resultado = pricer_forward(spot, dias, i_usd=sofr, params_curva_cop=params)
        sensibilidades = resultado.sensibilidades
        brecha = comparar_convenciones(spot, dias, i_usd=sofr, params_curva_cop=params)
        filas.append(
            {
                "plazo_dias": dias,
                "tasa_cop": resultado.i_cop,
                "forward": resultado.precio,
                "puntos_forward": resultado.puntos_forward,
                "devaluacion_ea": resultado.devaluacion_anualizada,
                "delta_spot": sensibilidades["delta_spot"],
                "dv01_cop": sensibilidades["dv01_cop"],
                "dv01_usd": sensibilidades["dv01_usd"],
                "theta_dia": sensibilidades["theta_dia"],
                "brecha_convenciones_bps": brecha["diferencia_bps"],
            }
        )
    _escribir_tabla(hoja, pd.DataFrame(filas).round(8))


def _hoja_validacion(wb: Workbook, params: NSSParams, spot: float, sofr: float) -> None:
    """Valores de referencia de Python junto a la fórmula VBA que debe reproducirlos.

    Es el control de paridad entre las dos implementaciones. La comparación se hace
    abriendo el libro: la columna ``resultado_vba`` se evalúa en Excel y debe coincidir
    con ``valor_python``.
    """
    hoja = wb.create_sheet("Validacion")
    cupon, plazo_bono, frecuencia = BONO_EJEMPLO
    flujos = flujos_bono(cupon, plazo_bono, frecuencia)

    casos: list[tuple[str, float, str]] = [
        (
            "Tasa cero cupón a 1 año",
            float(tasa_cero_cupon(1.0, params)),
            "=TASA_CERO_CUPON(1,NSS_PARAMS)",
        ),
        (
            "Tasa cero cupón a 5 años",
            float(tasa_cero_cupon(5.0, params)),
            "=TASA_CERO_CUPON(5,NSS_PARAMS)",
        ),
        (
            "Tasa cero cupón a 10 años",
            float(tasa_cero_cupon(10.0, params)),
            "=TASA_CERO_CUPON(10,NSS_PARAMS)",
        ),
        (
            "Factor de descuento a 5 años",
            float(factor_descuento(5.0, params)),
            "=FACTOR_DESCUENTO(5,NSS_PARAMS)",
        ),
        (
            "Forward 1a-5a",
            tasa_forward(1.0, 5.0, params),
            "=TASA_FORWARD(1,5,NSS_PARAMS)",
        ),
        (
            f"VP bono {cupon:.2%} {plazo_bono:.0f}a",
            valor_presente(flujos, params),
            f"=VP_TES({cupon},{plazo_bono:.0f},{frecuencia},100,NSS_PARAMS)",
        ),
        (
            f"DV01 bono {cupon:.2%} {plazo_bono:.0f}a",
            dv01(flujos, params),
            f"=DV01_TES({cupon},{plazo_bono:.0f},{frecuencia},100,NSS_PARAMS)",
        ),
        (
            f"Duración mod. bono {cupon:.2%}",
            duracion_modificada(flujos, params),
            f"=DURACION_MOD_TES({cupon},{plazo_bono:.0f},{frecuencia},100,NSS_PARAMS)",
        ),
    ]

    for dias in (90, 180, 360):
        resultado = pricer_forward(spot, dias, i_usd=sofr, params_curva_cop=params)
        casos.append(
            (
                f"Forward USD/COP {dias}d",
                resultado.precio,
                f"=FORWARD_USDCOP(TRM_SPOT,{dias},SOFR_ACTUAL,,NSS_PARAMS)",
            )
        )
        casos.append(
            (
                f"DV01 COP forward {dias}d",
                resultado.sensibilidades["dv01_cop"],
                f"=DV01_COP_FWD(TRM_SPOT,{dias},SOFR_ACTUAL,,NSS_PARAMS)",
            )
        )

    hoja["A1"] = "Paridad Python ↔ VBA"
    hoja["A1"].font = _TITULO
    hoja["A2"] = (
        "La columna 'resultado_vba' se evalúa en Excel al abrir el libro. "
        "Debe coincidir con 'valor_python'; 'diferencia' lo cuantifica."
    )

    encabezados = (
        "concepto",
        "valor_python",
        "resultado_vba",
        "diferencia",
        "formula_vba",
    )
    for j, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=4, column=j, value=texto)
        celda.font = _CABECERA
        celda.fill = _RELLENO

    for i, (concepto, valor, formula) in enumerate(casos, start=5):
        hoja.cell(row=i, column=1, value=concepto).border = _BORDE
        hoja.cell(row=i, column=2, value=float(valor)).number_format = "0.00000000"
        hoja.cell(row=i, column=3, value=formula).number_format = "0.00000000"
        hoja.cell(row=i, column=4, value=f"=C{i}-B{i}").number_format = "0.00E+00"
        # Copia textual de la fórmula, legible aunque Excel no haya recalculado.
        hoja.cell(row=i, column=5, value="'" + formula)

    for columna, ancho in zip("ABCDE", (34, 18, 18, 14, 52)):
        hoja.column_dimensions[columna].width = ancho


def exportar_libro(
    calibracion: ResultadoCalibracion,
    nodos: pd.DataFrame,
    spot: float,
    sofr: float,
    fecha_curva: date,
    ruta: Path | None = None,
) -> Path:
    """Genera el libro Excel con la curva calibrada y los rangos con nombre.

    Args:
        calibracion: Resultado de :func:`~motor_tes.curva_nss.calibrar_nss`.
        nodos: Nodos de mercado usados, con columna ``fuente`` si está disponible.
        spot: TRM contado, en pesos por dólar.
        sofr: SOFR vigente, en decimal.
        fecha_curva: Fecha de referencia de la curva.
        ruta: Destino del ``.xlsx``. Por defecto :data:`RUTA_LIBRO_BASE`.

    Returns:
        Ruta del libro escrito.
    """
    ruta = ruta or RUTA_LIBRO_BASE
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _hoja_parametros(wb, calibracion, spot, sofr, fecha_curva)
    _hoja_nodos(wb, calibracion, nodos)
    _hoja_curva(wb, calibracion.params)
    _hoja_forwards(wb, calibracion.params, spot, sofr)
    _hoja_validacion(wb, calibracion.params, spot, sofr)
    wb.save(ruta)
    return ruta
